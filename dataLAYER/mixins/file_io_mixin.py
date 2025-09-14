from __future__ import annotations

import os
import uuid
import csv
from typing import Optional

import pandas as pd  # type: ignore
from openpyxl import load_workbook  # type: ignore


class FileIOMixin:
    """提供健壮的文件读取与标准化能力。"""

    def _read_file_robustly(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        健壮地读取文件，特别是对 Excel 进行特殊处理。
        支持 .xlsx / .xls / .csv
        """
        if file_path.lower().endswith('.xlsx'):
            temp_csv_path = f"data/temp_{uuid.uuid4()}.csv"
            workbook = None
            try:
                # 1. 使用 openpyxl 加载工作簿
                workbook = load_workbook(filename=file_path, read_only=True)
                if not workbook.sheetnames:
                    raise ValueError("Excel文件中未找到任何工作表。")

                # 2. 选择第一个工作表
                sheet = workbook[workbook.sheetnames[0]]

                # 3. 手动将数据写入临时CSV文件
                with open(temp_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    for row in sheet.iter_rows(values_only=True):
                        writer.writerow([str(cell) if cell is not None else "" for cell in row])

                # 4. 关闭workbook以释放文件句柄
                workbook.close()
                workbook = None

                # 5. 让 pandas 读取这个干净的CSV文件
                data = pd.read_csv(temp_csv_path)
                print(f"成功通过 openpyxl->CSV 手动转换并读取了: {os.path.basename(file_path)}")
                return data

            except Exception as e:
                print(f"警告：使用 openpyxl 转换 '{os.path.basename(file_path)}' 失败。尝试使用本地Excel程序转换。详细信息: {e}")

                # 尝试第二种方法：使用 win32com 调用本地 Excel 转换
                if os.name == 'nt':  # 只在 Windows 系统上尝试
                    excel = None
                    wb = None
                    try:
                        import win32com.client as win32  # type: ignore
                        excel = win32.Dispatch('Excel.Application')
                        excel.Visible = False
                        excel.DisplayAlerts = False  # 关闭警告对话框
                        wb = excel.Workbooks.Open(os.path.abspath(file_path))

                        # 另存为CSV, 6代表xlCSV
                        wb.SaveAs(os.path.abspath(temp_csv_path), FileFormat=6)
                        wb.Close(False)
                        wb = None
                        excel.Application.Quit()
                        
                        # 确保COM对象被释放
                        import gc
                        del excel
                        gc.collect()
                        excel = None

                        data = pd.read_csv(temp_csv_path)
                        print(f"成功通过本地Excel->CSV转换并读取了: {os.path.basename(file_path)}")
                        return data
                    except Exception as com_e:
                        print(f"错误：使用本地Excel程序转换 '{os.path.basename(file_path)}' 也失败了。详细信息: {com_e}")
                        # 确保COM对象被清理
                        if wb:
                            try:
                                wb.Close(False)
                            except:
                                pass
                        if excel:
                            try:
                                excel.Application.Quit()
                                del excel
                                import gc
                                gc.collect()
                            except:
                                pass
                        return None
                else:
                    print("错误：非Windows系统，无法调用本地Excel程序。")
                    return None
            finally:
                # 确保workbook被关闭
                if workbook:
                    try:
                        workbook.close()
                    except:
                        pass
                
                # 删除临时文件
                if os.path.exists(temp_csv_path):
                    try:
                        os.remove(temp_csv_path)
                        print(f"已删除临时文件: {temp_csv_path}")
                    except Exception as e:
                        print(f"删除临时文件失败: {e}")

        elif file_path.lower().endswith('.xls'):
            temp_csv_path = None
            try:
                # 使用pandas读取.xls文件，并指定引擎确保文件正确关闭
                with pd.ExcelFile(file_path, engine='xlrd') as xls:
                    excel_df = pd.read_excel(xls)
                
                temp_csv_path = f"data/temp_{uuid.uuid4()}.csv"
                excel_df.to_csv(temp_csv_path, index=False)
                data = pd.read_csv(temp_csv_path)
                print(f"成功通过临时CSV标准化并读取了Excel文件: {os.path.basename(file_path)}")
                return data
            except Exception as e:
                print(f"错误：无法读取.xls文件 '{os.path.basename(file_path)}'。详细信息: {e}")
                return None
            finally:
                if temp_csv_path and os.path.exists(temp_csv_path):
                    try:
                        os.remove(temp_csv_path)
                        print(f"已删除临时文件: {temp_csv_path}")
                    except Exception as e:
                        print(f"删除临时文件失败: {e}")

        elif file_path.lower().endswith('.csv'):
            try:
                return pd.read_csv(file_path)
            except Exception as e:
                print(f"读取CSV文件时发生错误 '{os.path.basename(file_path)}': {e}")
                return None
        else:
            print(f"错误: 不支持的文件格式 {os.path.basename(file_path)}")
            return None


